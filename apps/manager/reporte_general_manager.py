from io import BytesIO
from typing import Any, Dict, List

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.api.config.exceptions.company_exception import NoCompanyAvailableException
from apps.mongo.daos.empresa_dao import EmpresaDAO
from apps.mongo.daos.periodo_contable_dao import PeriodoContableDAO
from apps.mongo.models.empresa import Empresa
from apps.mongo.models.extensions.balance_general import BalanceGeneral
from apps.mongo.models.extensions.estado_resultados import EstadoResultados
from apps.mongo.models.periodo_contable import PeriodoContable
from apps.tools.objectid import ObjectId


class ReporteGeneralManager:
    def __init__(self) -> None:
        self._periodo_dao = PeriodoContableDAO()
        self._empresa_dao = EmpresaDAO()

    async def get_reporte_final(
        self, id_empresa: ObjectId, anios: list[int]
    ) -> tuple[BytesIO, str]:
        """Genera reporte financiero completo con análisis vertical, horizontal, razones y tendencias"""
        filters: Dict = {
            "id_empresa": id_empresa,
            "anio": {"$in": anios},
        }

        empresa: Empresa | None = await self._empresa_dao.get_by_id(id_empresa)

        if not empresa:
            raise NoCompanyAvailableException(
                f"No hay proyecto disponible con el id: {id_empresa}"
            )

        periodos: list[PeriodoContable] = await self._periodo_dao.get_all(**filters)

        # Ordenar periodos por año
        periodos_ordenados: List[PeriodoContable] = sorted(
            periodos, key=lambda x: x.anio
        )

        estados_resultados: list[EstadoResultados] = [
            periodo.estado_resultado
            for periodo in periodos_ordenados
            if periodo.estado_resultado
        ]
        balances_generales: list[BalanceGeneral] = [
            periodo.balance_general
            for periodo in periodos_ordenados
            if periodo.balance_general
        ]

        # Generar todos los análisis
        df_analisis_vertical_balance = self._analisis_vertical_balance(
            balances_generales, anios
        )
        df_analisis_vertical_resultados = self._analisis_vertical_resultados(
            estados_resultados, anios
        )
        df_analisis_horizontal_balance = self._analisis_horizontal_balance(
            balances_generales, anios
        )
        df_analisis_horizontal_resultados = self._analisis_horizontal_resultados(
            estados_resultados, anios
        )
        df_razones_liquidez = self._analisis_razones_liquidez(periodos_ordenados)
        df_razones_actividad = self._analisis_razones_actividad(periodos_ordenados)
        df_razones_endeudamiento = self._analisis_razones_endeudamiento(
            periodos_ordenados
        )
        df_razones_rentabilidad = self._analisis_razones_rentabilidad(
            periodos_ordenados
        )
        # df_tendencias = self._analisis_tendencias(periodos_ordenados)
        # df_dashboard = self._crear_dashboard_data(periodos_ordenados)
        df_resumen_ejecutivo = self._crear_resumen_ejecutivo(periodos_ordenados)

        # Crear archivo Excel
        excel_buffer = self._crear_excel_reporte(
            df_analisis_vertical_balance,
            df_analisis_vertical_resultados,
            df_analisis_horizontal_balance,
            df_analisis_horizontal_resultados,
            df_razones_liquidez,
            df_razones_actividad,
            df_razones_endeudamiento,
            df_razones_rentabilidad,
            # df_tendencias,
            # df_dashboard,
            df_resumen_ejecutivo,
        )

        return (
            excel_buffer,
            f"Reporte Financiero de {empresa.nombre}.xlsx",
        )

    def _analisis_vertical_balance(
        self, balances: List[BalanceGeneral], anios: List[int]
    ) -> pd.DataFrame:
        """Análisis vertical del Balance General - cada partida como % del total de activos"""
        if not balances:
            return pd.DataFrame()

        data = []
        for i, balance in enumerate(balances):
            if i < len(anios):
                anio = anios[i]
                total_activo = balance.total_activo

                data.append(
                    {
                        "Concepto": "ACTIVO CIRCULANTE",
                        "Año": anio,
                        "Valor": balance.total_activo_circulante,
                        "Porcentaje": round(
                            (
                                (balance.total_activo_circulante / total_activo * 100)
                                if total_activo
                                else 0
                            ),
                            2,
                        ),
                    }
                )
                data.append(
                    {
                        "Concepto": "Efectivo y Equivalentes",
                        "Año": anio,
                        "Valor": balance.efectivo_equivalentes,
                        "Porcentaje": round(
                            (
                                (balance.efectivo_equivalentes / total_activo * 100)
                                if total_activo
                                else 0
                            ),
                            2,
                        ),
                    }
                )
                data.append(
                    {
                        "Concepto": "Cuentas por Cobrar",
                        "Año": anio,
                        "Valor": balance.cuentas_por_cobrar,
                        "Porcentaje": round(
                            (
                                (balance.cuentas_por_cobrar / total_activo * 100)
                                if total_activo
                                else 0
                            ),
                            2,
                        ),
                    }
                )
                data.append(
                    {
                        "Concepto": "Inventarios",
                        "Año": anio,
                        "Valor": balance.inventarios,
                        "Porcentaje": round(
                            (
                                (balance.inventarios / total_activo * 100)
                                if total_activo
                                else 0
                            ),
                            2,
                        ),
                    }
                )
                data.append(
                    {
                        "Concepto": "ACTIVO NO CIRCULANTE",
                        "Año": anio,
                        "Valor": balance.total_activo_no_circulante,
                        "Porcentaje": round(
                            (
                                (
                                    balance.total_activo_no_circulante
                                    / total_activo
                                    * 100
                                )
                                if total_activo
                                else 0
                            ),
                            2,
                        ),
                    }
                )
                data.append(
                    {
                        "Concepto": "Propiedades, Planta y Equipo",
                        "Año": anio,
                        "Valor": balance.propiedades_plantas_equipos,
                        "Porcentaje": round(
                            (
                                (
                                    balance.propiedades_plantas_equipos
                                    / total_activo
                                    * 100
                                )
                                if total_activo
                                else 0
                            ),
                            2,
                        ),
                    }
                )
                data.append(
                    {
                        "Concepto": "PASIVO CIRCULANTE",
                        "Año": anio,
                        "Valor": balance.total_pasivo_circulante,
                        "Porcentaje": round(
                            (
                                (balance.total_pasivo_circulante / total_activo * 100)
                                if total_activo
                                else 0
                            ),
                            2,
                        ),
                    }
                )
                data.append(
                    {
                        "Concepto": "Cuentas por Pagar",
                        "Año": anio,
                        "Valor": balance.cuentas_por_pagar,
                        "Porcentaje": round(
                            (
                                (balance.cuentas_por_pagar / total_activo * 100)
                                if total_activo
                                else 0
                            ),
                            2,
                        ),
                    }
                )
                data.append(
                    {
                        "Concepto": "PASIVO A LARGO PLAZO",
                        "Año": anio,
                        "Valor": balance.total_pasivo_a_largo_plazo,
                        "Porcentaje": round(
                            (
                                (
                                    balance.total_pasivo_a_largo_plazo
                                    / total_activo
                                    * 100
                                )
                                if total_activo
                                else 0
                            ),
                            2,
                        ),
                    }
                )
                data.append(
                    {
                        "Concepto": "CAPITAL CONTABLE",
                        "Año": anio,
                        "Valor": balance.capital_social_y_utilidades_retenidas,
                        "Porcentaje": round(
                            (
                                (
                                    balance.capital_social_y_utilidades_retenidas
                                    / total_activo
                                    * 100
                                )
                                if total_activo
                                else 0
                            ),
                            2,
                        ),
                    }
                )

        df = pd.DataFrame(data)
        return df.pivot(index="Concepto", columns="Año", values=["Valor", "Porcentaje"])

    def _analisis_vertical_resultados(
        self, estados: List[EstadoResultados], anios: List[int]
    ) -> pd.DataFrame:
        """Análisis vertical del Estado de Resultados - cada partida como % de ventas netas"""
        if not estados:
            return pd.DataFrame()

        data = []
        for i, estado in enumerate(estados):
            if i < len(anios):
                anio = anios[i]
                ventas_netas = estado.ventas_netas

                data.append(
                    {
                        "Concepto": "Ventas Netas",
                        "Año": anio,
                        "Valor": estado.ventas_netas,
                        "Porcentaje": 100.0,
                    }
                )
                data.append(
                    {
                        "Concepto": "Costo de Ventas",
                        "Año": anio,
                        "Valor": estado.costo_ventas,
                        "Porcentaje": round(
                            (
                                (estado.costo_ventas / ventas_netas * 100)
                                if ventas_netas
                                else 0
                            ),
                            2,
                        ),
                    }
                )
                data.append(
                    {
                        "Concepto": "Utilidad Bruta",
                        "Año": anio,
                        "Valor": estado.utilidad_bruta,
                        "Porcentaje": round(
                            (
                                (estado.utilidad_bruta / ventas_netas * 100)
                                if ventas_netas
                                else 0
                            ),
                            2,
                        ),
                    }
                )
                data.append(
                    {
                        "Concepto": "Gastos Operativos",
                        "Año": anio,
                        "Valor": estado.gastos_operativos,
                        "Porcentaje": round(
                            (
                                (estado.gastos_operativos / ventas_netas * 100)
                                if ventas_netas
                                else 0
                            ),
                            2,
                        ),
                    }
                )
                data.append(
                    {
                        "Concepto": "Utilidad Operativa",
                        "Año": anio,
                        "Valor": estado.utilidad_operativa,
                        "Porcentaje": round(
                            (
                                (estado.utilidad_operativa / ventas_netas * 100)
                                if ventas_netas
                                else 0
                            ),
                            2,
                        ),
                    }
                )
                data.append(
                    {
                        "Concepto": "Resultado Financiero",
                        "Año": anio,
                        "Valor": estado.resultado_financieros,
                        "Porcentaje": round(
                            (
                                (estado.resultado_financieros / ventas_netas * 100)
                                if ventas_netas
                                else 0
                            ),
                            2,
                        ),
                    }
                )
                data.append(
                    {
                        "Concepto": "Utilidad Antes de Impuestos",
                        "Año": anio,
                        "Valor": estado.utilidad_ante_impuestos,
                        "Porcentaje": round(
                            (
                                (estado.utilidad_ante_impuestos / ventas_netas * 100)
                                if ventas_netas
                                else 0
                            ),
                            2,
                        ),
                    }
                )
                data.append(
                    {
                        "Concepto": "Impuesto sobre Utilidad",
                        "Año": anio,
                        "Valor": estado.impuesto_utilidad,
                        "Porcentaje": round(
                            (
                                (estado.impuesto_utilidad / ventas_netas * 100)
                                if ventas_netas
                                else 0
                            ),
                            2,
                        ),
                    }
                )
                data.append(
                    {
                        "Concepto": "Utilidad Neta",
                        "Año": anio,
                        "Valor": estado.utilidad_neta,
                        "Porcentaje": round(
                            (
                                (estado.utilidad_neta / ventas_netas * 100)
                                if ventas_netas
                                else 0
                            ),
                            2,
                        ),
                    }
                )

        df = pd.DataFrame(data)
        return df.pivot(index="Concepto", columns="Año", values=["Valor", "Porcentaje"])

    def _analisis_horizontal_balance(
        self, balances: List[BalanceGeneral], anios: List[int]
    ) -> pd.DataFrame:
        """Análisis horizontal del Balance General - variación entre periodos"""
        if len(balances) < 2:
            return pd.DataFrame()

        data = []
        conceptos = [
            ("Efectivo y Equivalentes", "efectivo_equivalentes"),
            ("Cuentas por Cobrar", "cuentas_por_cobrar"),
            ("Inventarios", "inventarios"),
            ("Total Activo Circulante", "total_activo_circulante"),
            ("Propiedades, Planta y Equipo", "propiedades_plantas_equipos"),
            ("Total Activo", "total_activo"),
            ("Cuentas por Pagar", "cuentas_por_pagar"),
            ("Total Pasivo Circulante", "total_pasivo_circulante"),
            ("Total Pasivo a Largo Plazo", "total_pasivo_a_largo_plazo"),
            ("Total Pasivo", "total_pasivo"),
            ("Capital Contable", "capital_social_y_utilidades_retenidas"),
        ]

        for concepto_nombre, campo in conceptos:
            row = {"Concepto": concepto_nombre}

            for i in range(len(balances)):
                if i < len(anios):
                    valor_actual = getattr(balances[i], campo)
                    row[f"{anios[i]}"] = valor_actual

                    if i > 0:  # Calcular variación respecto al año anterior
                        valor_anterior = getattr(balances[i - 1], campo)
                        variacion_absoluta = valor_actual - valor_anterior
                        variacion_porcentual = round(
                            (
                                (variacion_absoluta / valor_anterior * 100)
                                if valor_anterior != 0
                                else 0
                            ),
                            2,
                        )

                        row[f"Var. {anios[i-1]}-{anios[i]} ($)"] = variacion_absoluta
                        row[f"Var. {anios[i-1]}-{anios[i]} (%)"] = str(
                            variacion_porcentual
                        )

            data.append(row)

        return pd.DataFrame(data)

    def _analisis_horizontal_resultados(
        self, estados: List[EstadoResultados], anios: List[int]
    ) -> pd.DataFrame:
        """Análisis horizontal del Estado de Resultados - variación entre periodos"""
        if len(estados) < 2:
            return pd.DataFrame()

        data = []
        conceptos = [
            ("Ventas Netas", "ventas_netas"),
            ("Costo de Ventas", "costo_ventas"),
            ("Utilidad Bruta", "utilidad_bruta"),
            ("Gastos Operativos", "gastos_operativos"),
            ("Utilidad Operativa", "utilidad_operativa"),
            ("Resultado Financiero", "resultado_financieros"),
            ("Utilidad Antes de Impuestos", "utilidad_ante_impuestos"),
            ("Impuesto sobre Utilidad", "impuesto_utilidad"),
            ("Utilidad Neta", "utilidad_neta"),
        ]

        for concepto_nombre, campo in conceptos:
            row = {"Concepto": concepto_nombre}

            for i in range(len(estados)):
                if i < len(anios):
                    valor_actual = getattr(estados[i], campo)
                    row[f"{anios[i]}"] = valor_actual

                    if i > 0:  # Calcular variación respecto al año anterior
                        valor_anterior = getattr(estados[i - 1], campo)
                        variacion_absoluta = valor_actual - valor_anterior
                        variacion_porcentual = round(
                            (
                                (variacion_absoluta / valor_anterior * 100)
                                if valor_anterior != 0
                                else 0
                            ),
                            2,
                        )

                        row[f"Var. {anios[i-1]}-{anios[i]} ($)"] = variacion_absoluta
                        row[f"Var. {anios[i-1]}-{anios[i]} (%)"] = str(
                            variacion_porcentual
                        )

            data.append(row)

        return pd.DataFrame(data)

    def _analisis_razones_liquidez(
        self, periodos: List[PeriodoContable]
    ) -> pd.DataFrame:
        """Análisis de razones de liquidez"""
        data = []
        for periodo in periodos:
            if periodo.balance_general:
                data.append(
                    {
                        "Año": periodo.anio,
                        "Razón Corriente": round(
                            periodo.balance_general.razon_corriente, 2
                        ),
                        "Prueba Ácida": round(periodo.balance_general.prueba_acida, 2),
                        "Capital Neto de Trabajo": round(
                            periodo.balance_general.capital_neto_trabajo, 2
                        ),
                    }
                )

        return pd.DataFrame(data)

    def _analisis_razones_actividad(
        self, periodos: List[PeriodoContable]
    ) -> pd.DataFrame:
        """Análisis de razones de actividad"""
        data = []
        for periodo in periodos:
            data.append(
                {
                    "Año": periodo.anio,
                    "Rotación de Inventarios": round(periodo.rotacion_inventarios, 2),
                    "Días de Inventario": round(periodo.dias_inventario, 2),
                    "Rotación Ctas. por Cobrar": round(
                        periodo.rotacion_cuentas_cobrar, 2
                    ),
                    "Período Promedio de Cobro": round(
                        periodo.periodo_promedio_cobro, 2
                    ),
                    "Rotación Ctas. por Pagar": round(
                        periodo.rotacion_cuentas_pagar, 2
                    ),
                    "Período Promedio de Pago": round(periodo.periodo_promedio_pago, 2),
                    "Ciclo de Efectivo": round(periodo.ciclo_efectivo, 2),
                }
            )

        return pd.DataFrame(data)

    def _analisis_razones_endeudamiento(
        self, periodos: List[PeriodoContable]
    ) -> pd.DataFrame:
        """Análisis de razones de endeudamiento"""
        data = []
        for periodo in periodos:
            if periodo.balance_general:
                data.append(
                    {
                        "Año": periodo.anio,
                        "Endeudamiento Total (%)": round(
                            periodo.balance_general.endeudamiento_total, 2
                        ),
                        "Endeudamiento Patrimonial (%)": round(
                            periodo.balance_general.endeudamiento_patrimonial, 2
                        ),
                        "Apalancamiento Financiero": round(
                            periodo.balance_general.apalancamiento_financiero, 2
                        ),
                        "Cobertura de Intereses": (
                            round(periodo.cobertura_intereses, 2)
                            if periodo.cobertura_intereses != float("inf")
                            else "N/A"
                        ),
                    }
                )

        return pd.DataFrame(data)

    def _analisis_razones_rentabilidad(
        self, periodos: List[PeriodoContable]
    ) -> pd.DataFrame:
        """Análisis de razones de rentabilidad"""
        data = []
        for periodo in periodos:
            if periodo.estado_resultado:
                data.append(
                    {
                        "Año": periodo.anio,
                        "Margen Bruto (%)": round(
                            periodo.estado_resultado.margen_bruto, 2
                        ),
                        "Margen Operativo (%)": round(
                            periodo.estado_resultado.margen_operativo, 2
                        ),
                        "Margen Neto (%)": round(
                            periodo.estado_resultado.margen_neto, 2
                        ),
                        "ROA (%)": round(periodo.roa, 2),
                        "ROE (%)": round(periodo.roe, 2),
                    }
                )

        return pd.DataFrame(data)

    def _analisis_tendencias(self, periodos: List[PeriodoContable]) -> pd.DataFrame:
        """Análisis de tendencias de indicadores clave"""
        data = []
        for periodo in periodos:
            row: dict[Any, Any] = {"Año": periodo.anio}

            if periodo.estado_resultado:
                row.update(
                    {
                        "Ventas Netas": round(periodo.estado_resultado.ventas_netas, 2),
                        "Utilidad Bruta": round(
                            periodo.estado_resultado.utilidad_bruta, 2
                        ),
                        "Utilidad Operativa": round(
                            periodo.estado_resultado.utilidad_operativa, 2
                        ),
                        "Utilidad Neta": round(
                            periodo.estado_resultado.utilidad_neta, 2
                        ),
                        "Margen Bruto (%)": round(
                            periodo.estado_resultado.margen_bruto, 2
                        ),
                        "Margen Operativo (%)": round(
                            periodo.estado_resultado.margen_operativo, 2
                        ),
                        "Margen Neto (%)": round(
                            periodo.estado_resultado.margen_neto, 2
                        ),
                    }
                )  # type: ignore

            if periodo.balance_general:
                row.update(
                    {
                        "Total Activo": round(periodo.balance_general.total_activo, 2),
                        "Total Pasivo": round(periodo.balance_general.total_pasivo, 2),
                        "Capital Contable": round(
                            periodo.balance_general.capital_social_y_utilidades_retenidas,
                            2,
                        ),
                        "Razón Corriente": round(
                            periodo.balance_general.razon_corriente, 2
                        ),
                        "Endeudamiento Total (%)": round(
                            periodo.balance_general.endeudamiento_total, 2
                        ),
                    }
                )  # type: ignore

            row.update(
                {
                    "ROA (%)": round(periodo.roa, 2),
                    "ROE (%)": round(periodo.roe, 2),
                    "Rotación de Inventarios": round(periodo.rotacion_inventarios, 2),
                }
            )  # type: ignore

            data.append(row)

        return pd.DataFrame(data)

    def _crear_dashboard_data(self, periodos: List[PeriodoContable]) -> pd.DataFrame:
        """Crea datos consolidados para dashboard con KPIs principales"""
        data = []
        for periodo in periodos:
            row = {
                "Año": periodo.anio,
                "Período": f"{periodo.fecha_inicio} - {periodo.fecha_fin}",
            }

            # KPIs Financieros Principales
            if periodo.estado_resultado:
                row.update(
                    {
                        "Ventas Netas": round(periodo.estado_resultado.ventas_netas, 2),
                        "EBITDA": round(
                            periodo.estado_resultado.utilidad_operativa, 2
                        ),  # Simplificado
                        "Utilidad Neta": round(
                            periodo.estado_resultado.utilidad_neta, 2
                        ),
                        "Margen Bruto (%)": round(
                            periodo.estado_resultado.margen_bruto, 2
                        ),
                        "Margen Operativo (%)": round(
                            periodo.estado_resultado.margen_operativo, 2
                        ),
                        "Margen Neto (%)": round(
                            periodo.estado_resultado.margen_neto, 2
                        ),
                    }
                )
            else:
                row.update(
                    {
                        "Ventas Netas": 0,
                        "EBITDA": 0,
                        "Utilidad Neta": 0,
                        "Margen Bruto (%)": 0,
                        "Margen Operativo (%)": 0,
                        "Margen Neto (%)": 0,
                    }
                )

            if periodo.balance_general:
                row.update(
                    {
                        "Total Activos": round(periodo.balance_general.total_activo, 2),
                        "Total Pasivos": round(periodo.balance_general.total_pasivo, 2),
                        "Patrimonio": round(
                            periodo.balance_general.capital_social_y_utilidades_retenidas,
                            2,
                        ),
                        "Activo Circulante": round(
                            periodo.balance_general.total_activo_circulante, 2
                        ),
                        "Pasivo Circulante": round(
                            periodo.balance_general.total_pasivo_circulante, 2
                        ),
                        "Efectivo": round(
                            periodo.balance_general.efectivo_equivalentes, 2
                        ),
                        "Inventarios": round(periodo.balance_general.inventarios, 2),
                        "Cuentas por Cobrar": round(
                            periodo.balance_general.cuentas_por_cobrar, 2
                        ),
                    }
                )
            else:
                row.update(
                    {
                        "Total Activos": 0,
                        "Total Pasivos": 0,
                        "Patrimonio": 0,
                        "Activo Circulante": 0,
                        "Pasivo Circulante": 0,
                        "Efectivo": 0,
                        "Inventarios": 0,
                        "Cuentas por Cobrar": 0,
                    }
                )

            # Ratios Clave
            row.update(
                {
                    "Liquidez Corriente": (
                        round(periodo.balance_general.razon_corriente, 2)
                        if periodo.balance_general
                        else 0
                    ),
                    "Prueba Ácida": (
                        round(periodo.balance_general.prueba_acida, 2)
                        if periodo.balance_general
                        else 0
                    ),
                    "Endeudamiento (%)": (
                        round(periodo.balance_general.endeudamiento_total, 2)
                        if periodo.balance_general
                        else 0
                    ),
                    "ROA (%)": round(periodo.roa, 2),
                    "ROE (%)": round(periodo.roe, 2),
                    "Rotación Inventarios": round(periodo.rotacion_inventarios, 2),
                    "Días de Cobro": round(periodo.periodo_promedio_cobro, 0),
                    "Cobertura Intereses": (
                        round(periodo.cobertura_intereses, 2)
                        if periodo.cobertura_intereses != float("inf")
                        else "N/A"
                    ),
                }
            )

            # Indicadores de Crecimiento (si hay período anterior)
            if len(data) > 0:  # Hay período anterior
                periodo_anterior = data[-1]

                # Crecimiento de Ventas
                ventas_anterior = periodo_anterior.get("Ventas Netas", 0)
                if ventas_anterior > 0:
                    crecimiento_ventas = (
                        (row["Ventas Netas"] - ventas_anterior) / ventas_anterior
                    ) * 100
                    row["Crecimiento Ventas (%)"] = round(crecimiento_ventas, 2)
                else:
                    row["Crecimiento Ventas (%)"] = 0

                # Crecimiento de Activos
                activos_anterior = periodo_anterior.get("Total Activos", 0)
                if activos_anterior > 0:
                    crecimiento_activos = (
                        (row["Total Activos"] - activos_anterior) / activos_anterior
                    ) * 100
                    row["Crecimiento Activos (%)"] = round(crecimiento_activos, 2)
                else:
                    row["Crecimiento Activos (%)"] = 0
            else:
                row["Crecimiento Ventas (%)"] = 0
                row["Crecimiento Activos (%)"] = 0

            # Semáforo de Alertas (indicadores de salud financiera)
            alertas = []

            # Alerta de Liquidez
            if row["Liquidez Corriente"] < 1.0:
                alertas.append("LIQUIDEZ BAJA")
            elif row["Liquidez Corriente"] > 3.0:
                alertas.append("EXCESO LIQUIDEZ")

            # Alerta de Endeudamiento
            if row["Endeudamiento (%)"] > 70:
                alertas.append("ALTO ENDEUDAMIENTO")

            # Alerta de Rentabilidad
            if row["Margen Neto (%)"] < 0:
                alertas.append("PÉRDIDAS")
            elif row["Margen Neto (%)"] < 5:
                alertas.append("BAJA RENTABILIDAD")

            # Alerta de Rotación
            if row["Rotación Inventarios"] < 2:
                alertas.append("INVENTARIO LENTO")

            row["Alertas"] = "; ".join(alertas) if alertas else "SALUDABLE"

            data.append(row)

        return pd.DataFrame(data)

    def _crear_resumen_ejecutivo(self, periodos: List[PeriodoContable]) -> pd.DataFrame:
        """Crea resumen ejecutivo con métricas clave"""
        if not periodos:
            return pd.DataFrame()

        # Obtener último período
        ultimo_periodo = periodos[-1]

        # Obtener período anterior si existe
        periodo_anterior = periodos[-2] if len(periodos) > 1 else None

        metricas = []

        # Sección: Resultados Financieros
        metricas.append(
            {
                "Categoría": "RESULTADOS FINANCIEROS",
                "Métrica": "Ventas Netas",
                "Valor Actual": (
                    ultimo_periodo.estado_resultado.ventas_netas
                    if ultimo_periodo.estado_resultado
                    else 0
                ),
                "Período Anterior": (
                    periodo_anterior.estado_resultado.ventas_netas
                    if periodo_anterior and periodo_anterior.estado_resultado
                    else 0
                ),
                "Variación (%)": 0,
                "Estado": "INFO",
            }
        )

        metricas.append(
            {
                "Categoría": "RESULTADOS FINANCIEROS",
                "Métrica": "Utilidad Neta",
                "Valor Actual": (
                    ultimo_periodo.estado_resultado.utilidad_neta
                    if ultimo_periodo.estado_resultado
                    else 0
                ),
                "Período Anterior": (
                    periodo_anterior.estado_resultado.utilidad_neta
                    if periodo_anterior and periodo_anterior.estado_resultado
                    else 0
                ),
                "Variación (%)": 0,
                "Estado": (
                    "POSITIVO"
                    if (
                        ultimo_periodo.estado_resultado.utilidad_neta
                        if ultimo_periodo.estado_resultado
                        else 0
                    )
                    > 0
                    else "NEGATIVO"
                ),
            }
        )

        # Sección: Posición Financiera
        metricas.append(
            {
                "Categoría": "POSICIÓN FINANCIERA",
                "Métrica": "Total Activos",
                "Valor Actual": (
                    ultimo_periodo.balance_general.total_activo
                    if ultimo_periodo.balance_general
                    else 0
                ),
                "Período Anterior": (
                    periodo_anterior.balance_general.total_activo
                    if periodo_anterior and periodo_anterior.balance_general
                    else 0
                ),
                "Variación (%)": 0,
                "Estado": "INFO",
            }
        )

        metricas.append(
            {
                "Categoría": "POSICIÓN FINANCIERA",
                "Métrica": "Efectivo y Equivalentes",
                "Valor Actual": (
                    ultimo_periodo.balance_general.efectivo_equivalentes
                    if ultimo_periodo.balance_general
                    else 0
                ),
                "Período Anterior": (
                    periodo_anterior.balance_general.efectivo_equivalentes
                    if periodo_anterior and periodo_anterior.balance_general
                    else 0
                ),
                "Variación (%)": 0,
                "Estado": "INFO",
            }
        )

        # Sección: Ratios Clave
        metricas.append(
            {
                "Categoría": "RATIOS CLAVE",
                "Métrica": "Liquidez Corriente",
                "Valor Actual": (
                    ultimo_periodo.balance_general.razon_corriente
                    if ultimo_periodo.balance_general
                    else 0
                ),
                "Período Anterior": (
                    periodo_anterior.balance_general.razon_corriente
                    if periodo_anterior and periodo_anterior.balance_general
                    else 0
                ),
                "Variación (%)": 0,
                "Estado": (
                    "POSITIVO"
                    if (
                        ultimo_periodo.balance_general.razon_corriente
                        if ultimo_periodo.balance_general
                        else 0
                    )
                    >= 1.2
                    else "ALERTA"
                ),
            }
        )

        metricas.append(
            {
                "Categoría": "RATIOS CLAVE",
                "Métrica": "ROE (%)",
                "Valor Actual": f"{ultimo_periodo.roe:.2f}%",
                "Período Anterior": (
                    f"{periodo_anterior.roe:.2f}%" if periodo_anterior else "0.00%"
                ),
                "Variación (%)": 0,
                "Estado": (
                    "POSITIVO"
                    if ultimo_periodo.roe > 15
                    else "ALERTA" if ultimo_periodo.roe > 0 else "NEGATIVO"
                ),
            }
        )

        # Calcular variaciones
        for metrica in metricas:
            if isinstance(metrica["Período Anterior"], str) and "%" in str(
                metrica["Período Anterior"]
            ):
                # Para métricas que ya son porcentajes, mantener el formato
                continue
            elif metrica["Período Anterior"] != 0:
                variacion = (
                    (metrica["Valor Actual"] - metrica["Período Anterior"])
                    / metrica["Período Anterior"]
                ) * 100
                metrica["Variación (%)"] = f"{variacion:.2f}%"
            else:
                metrica["Variación (%)"] = "0.00%"

        return pd.DataFrame(metricas)

    def _crear_excel_reporte(
        self,
        df_vert_balance,
        df_vert_resultados,
        df_horiz_balance,
        df_horiz_resultados,
        df_liquidez,
        df_actividad,
        df_endeudamiento,
        df_rentabilidad,
        # df_tendencias,
        # df_dashboard,
        df_resumen,
    ) -> BytesIO:
        """Crea archivo Excel con todas las hojas del análisis financiero"""

        excel_buffer = BytesIO()

        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            # Hoja de Resumen Ejecutivo (primera hoja)
            if not df_resumen.empty:
                df_resumen.to_excel(
                    writer, sheet_name="📊 Resumen Ejecutivo", index=False
                )

            # Hoja de Dashboard
            # if not df_dashboard.empty:
            #     df_dashboard.to_excel(
            #         writer, sheet_name="🎯 Dashboard KPIs", index=False
            #     )

            # Escribir cada DataFrame en una hoja diferente
            if not df_vert_balance.empty:
                df_vert_balance.to_excel(writer, sheet_name="📈 Vert Balance")

            if not df_horiz_balance.empty:
                df_horiz_balance.to_excel(
                    writer, sheet_name="📊 Horiz Balance", index=False
                )

            if not df_vert_resultados.empty:
                df_vert_resultados.to_excel(writer, sheet_name="📈 Vert Resultados")

            if not df_horiz_resultados.empty:
                df_horiz_resultados.to_excel(
                    writer, sheet_name="📊 Horiz Resultados", index=False
                )

            if not df_liquidez.empty:
                df_liquidez.to_excel(writer, sheet_name="💧 Liquidez", index=False)

            if not df_actividad.empty:
                df_actividad.to_excel(writer, sheet_name="🔄 Actividad", index=False)

            if not df_endeudamiento.empty:
                df_endeudamiento.to_excel(
                    writer, sheet_name="💳 Endeudamiento", index=False
                )

            if not df_rentabilidad.empty:
                df_rentabilidad.to_excel(
                    writer, sheet_name="💰 Rentabilidad", index=False
                )

            # if not df_tendencias.empty:
            #     df_tendencias.to_excel(writer, sheet_name="📈 Tendencias", index=False)

        # Aplicar formato al archivo Excel
        excel_buffer.seek(0)
        workbook = openpyxl.load_workbook(excel_buffer)

        # Formato para encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Formato especial para Dashboard y Resumen Ejecutivo
        kpi_font = Font(bold=True, size=12)
        alert_font = Font(bold=True, color="FF0000")  # Rojo para alertas
        positive_font = Font(bold=True, color="008000")  # Verde para positivos

        # Formato para porcentajes
        percentage_format = "0.00%"

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # Aplicar formato a la primera fila (encabezados)
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Aplicar formato de porcentajes a columnas específicas
            if "Análisis Vertical" in sheet_name:
                # Para análisis vertical, las columnas de porcentaje suelen estar después de las de valores
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        if (
                            cell.value
                            and isinstance(cell.value, (int, float))
                            and abs(cell.value) <= 1
                        ):
                            # Si es un decimal menor a 1, probablemente es un porcentaje en formato decimal
                            cell.number_format = percentage_format
                        elif isinstance(cell.value, str) and "%" in str(cell.value):
                            # Si ya tiene el símbolo %, aplicar formato de porcentaje
                            try:
                                numeric_value = (
                                    float(str(cell.value).replace("%", "")) / 100
                                )
                                cell.value = numeric_value  # type: ignore
                                cell.number_format = percentage_format
                            except ValueError:
                                pass

            elif "Razones" in sheet_name:
                # Para hojas de razones financieras, identificar columnas de porcentajes
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        if isinstance(cell.value, str) and "%" in str(cell.value):
                            try:
                                numeric_value = (
                                    float(str(cell.value).replace("%", "")) / 100
                                )
                                cell.value = numeric_value  # type: ignore
                                cell.number_format = percentage_format
                            except ValueError:
                                pass

            elif "Dashboard" in sheet_name:
                # Aplicar formato de porcentajes a columnas específicas del dashboard
                for row in worksheet.iter_rows(min_row=2):
                    # Columnas que contienen porcentajes (márgenes, crecimientos, etc.)
                    percentage_columns = [
                        6,
                        7,
                        8,
                        19,
                        20,
                    ]  # Ajustar según la estructura real

                    for col_idx in percentage_columns:
                        if len(row) > col_idx:
                            cell = row[col_idx]
                            if isinstance(cell.value, str) and "%" in str(cell.value):
                                try:
                                    numeric_value = (
                                        float(str(cell.value).replace("%", "")) / 100
                                    )
                                    cell.value = numeric_value  # type: ignore
                                    cell.number_format = percentage_format
                                except ValueError:
                                    pass
                            elif (
                                isinstance(cell.value, (int, float))
                                and abs(cell.value) <= 1
                            ):
                                cell.number_format = percentage_format

                    # Resaltar KPIs importantes
                    if len(row) > 15 and row[15].value and row[15].value != "SALUDABLE":
                        row[15].font = alert_font

                    # Márgenes negativos en rojo
                    for col_idx in [6, 7, 8]:  # Márgenes
                        if len(row) > col_idx:
                            cell_value = row[col_idx].value
                            if (
                                cell_value is not None
                                and isinstance(cell_value, (int, float))
                                and cell_value < 0
                            ):
                                row[col_idx].font = alert_font

                    # Crecimiento positivo en verde
                    for col_idx in [19, 20]:  # Crecimientos
                        if len(row) > col_idx:
                            cell_value = row[col_idx].value
                            if (
                                cell_value is not None
                                and isinstance(cell_value, (int, float))
                                and cell_value > 0
                            ):
                                row[col_idx].font = positive_font

            elif "Resumen" in sheet_name:
                # Formato especial para Resumen Ejecutivo
                for row in worksheet.iter_rows(min_row=2):
                    # Aplicar formato de porcentaje a la columna "Variación (%)"
                    if (
                        len(row) > 4
                    ):  # Asumiendo que "Variación (%)" está en la columna 5 (índice 4)
                        cell = row[4]
                        if isinstance(cell.value, str) and "%" in str(cell.value):
                            try:
                                numeric_value = (
                                    float(str(cell.value).replace("%", "")) / 100
                                )
                                cell.value = numeric_value  # type: ignore
                                cell.number_format = percentage_format
                            except ValueError:
                                pass

                    if len(row) > 5:
                        estado = row[5].value
                        if estado == "NEGATIVO" or estado == "ALERTA":
                            for cell in row:
                                cell.font = alert_font
                        elif estado == "POSITIVO":
                            for cell in row:
                                cell.font = positive_font

            elif "Tendencias" in sheet_name:
                # Para análisis de tendencias, aplicar formato a columnas de crecimiento
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        if isinstance(cell.value, str) and "%" in str(cell.value):
                            try:
                                numeric_value = (
                                    float(str(cell.value).replace("%", "")) / 100
                                )
                                cell.value = numeric_value  # type: ignore
                                cell.number_format = percentage_format
                            except ValueError:
                                pass

            # Ajustar ancho de columnas
            for idx, column in enumerate(worksheet.columns, 1):
                max_length = 0
                column_letter = get_column_letter(idx)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass

                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer
